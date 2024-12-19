import os
from docxtpl import DocxTemplate
from students import Students
from teachers import Teachers
from groups import Groups
from schedule import Schedule
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import date

class Managers:
    """Класс, имеющий методы для управления системой и данными"""
    def print_file(self):
        """Функция открывает соответствущий документ (анкету).
        Используется модуль os - позволяет работать с ОС, файлами и тд"""
        f = input("Выберите анкету для печати: \n 1. Для студентов\n 2. Для преподавателей\n 3. Для родителей\n ")
        if f == '1':
            os.startfile('students_form.docx')
        elif f == '2':
            os.startfile('teachers_form.docx')
        elif f == '3':
            os.startfile('parents_form.docx')
        else:
            print("Введите соответсвующий номер анкеты")

    def register_student(self):
        """Функция для регистрации студента.
        Вводятся данные из которых создается экземпляр класса. Затем записываются в файл
        students через запятую. Также логин сохраняется в файле users (пароль - не вводится менеджером)"""
        user_id = input("Введите логин студента: ")
        group = input("Введите группу студента: ")
        surname = input("Введите фамилию: ")
        name = input("Введите имя: ")
        last_name = input("Введите отчество: ")
        e_mail = input("Введите e-mail: ")
        phone = input("Введите телефон: ")
        birth_date = input("Введите дату рождения (ДД.ММ.ГГГГ): ")
        student = Students(user_id, group, surname, name, last_name, e_mail, phone, birth_date)
        with open('students.txt', 'a', encoding='utf-8') as file:
            file.write(
                f"{student.user_id},{student.group},{student.surname},{student.name},{student.last_name},{student.e_mail},{student.phone},{student.birth_date}\n")
        with open('users.txt', 'a', encoding='utf-8') as file:
            file.write(
                f"{student.user_id},_,студент\n")

    def register_teacher(self):
        """Функция аналогичная предыдущему, регистрирует нового преподавателя.
        Данные сохраняются в файлах users и teachers"""
        try:
            teacher_id = input("Введите логин преподавателя: ")
            surname = input("Введите фамилию: ")
            name = input("Введите имя: ")
            last_name = input("Введите отчество: ")
            e_mail = input("Введите e-mail: ")
            phone = input("Введите телефон: ")
            birth_date = input("Введите день рождения преподавателя: ")
            teacher = Teachers(teacher_id, surname, name, last_name, e_mail, phone, birth_date)
            with open('teachers.txt', 'a', encoding='utf-8') as file:
                file.write(
                    f"{teacher.user_id},{teacher.surname},{teacher.name},{teacher.last_name},{teacher.e_mail},{teacher.phone},{birth_date}\n")
            with open('users.txt', 'a', encoding='utf-8') as file:
                file.write(
                    f"{teacher.user_id},_,преподаватель\n")
            print("Пользователь успешно зарегистрирован")
        except Exception:
            print("Ошибка. Введите корректные данные")

    def get_schedule_for_group(self):
        """Функция для вывода расписания для определенной группы.
        Вводится номер группы, далее при чтении файла с расписанием проверяется равенство нулевого элемента строки с номером группы
        При наличии такой строки, создается экземпляр класса с данными из файла и выводятся с помощью метода этого класса"""
        group = input("Введите группу: ")
        with open('schedule.txt', 'r', encoding='utf') as file:
            for line in file:
                sh = line.strip().split(',')
                if group == sh[0]:
                    schedule = Schedule(sh[0], sh[1], sh[2], sh[3], sh[4])
                    print(schedule)

    def get_info_student(self):
        """Функция для вывода информации о студенте. Вводится ФИ студента, сохраняется в переменной student_name
        При чтении файла сравнивается имя и фамилия отдельно с каждым элементом строки из файла.
        При наличи такой записи, создается экземпляр и выводятся данные"""
        student_name = input("Введите ФИ студента: ").split()
        with open('students.txt', 'r', encoding='utf-8') as file:
            for line in file:
                st = line.strip().split(',')
                if len(st) == 8:
                    if st[2] == student_name[0] and st[3] == student_name[1]:
                        student = Students(st[0], st[1], st[2], st[3], st[4], st[5], st[6], st[7])
                        print(student)
            else:
                print("Студент с таким именем не найден")

    def get_info_teacher(self):
        """Функция аналогична предыдущему. Выводит полную информацию о педагоге по его ФИ"""
        teacher_name = input("Введите ФИ преподавателя: ").split()
        with open('teachers.txt', 'r', encoding='utf-8') as file:
            for line in file:
                tch = line.strip().split(',')
                if tch[1] == teacher_name[0] and tch[2] == teacher_name[1]:
                    teacher = Teachers(tch[0], tch[1], tch[2], tch[3], tch[4], tch[5], tch[6])
                    print(teacher)
            else:
                print("Преподаватель с таким именем не найден")

    def get_info(self):
        """Функция для упрщения процесса получения информации о пользователе"""
        while True:
            f = input("Выберите следующее действие:\n 1. Личные данные студента \n 2. Личные данные преподавателя\n")
            if f == '1':
                self.get_info_student()
            elif f == '2':
                self.get_info_teacher()
            else:
                print("Ошибка. Введите номер соответствующего действия.")

    def print_certificate(self, student_name, language, level, course, organization, date):
        """Функция для создания именованного сертификата студенту, окончившему курс.
         Данные, вводимые с клавиатуры, сохраняются в словаре и создается файл
         Аргументы:
            student_name - ФИО студента
            language - язык, который был изучен
            level уровень владения языком на момент завершения курса
            course - название курса
            organization - название организации
            date - дата выдачи сертификата"""
        try:
            doc = DocxTemplate("certificate_templ.docx")
            context = {
                'student_name': student_name,
                'language': language,
                'level': level,
                'course': course,
                'organization': organization,
                'date': date}
            doc.render(context)
            file = f"certificate_{student_name.replace(' ', '_')}.docx"
            doc.save(file)
            os.startfile(file)
        except Exception as e:
            print(f"Ошибка {e}")

    def new_lesson(self):
        """Функция для добавления нового занятия в расписание.
        Данные вносятся и сохраняются в экземпляре класса, затем происходит запись этих данных в файл"""
        try:
            group = input("Введите группу: ")
            day = input("Введите день недели: ")
            start = input("Введите начало занятия: ")
            course = input("Введите курс: ")
            teacher = input("Введите преподавателя: ")
            new_schedule = Schedule(group, day, start, course, teacher)
            with open('schedule.txt', 'a', encoding='utf-8') as file:
                file.write(
                    f"{new_schedule.group},{new_schedule.day_of_week},{new_schedule.start_time},{new_schedule.course},{new_schedule.teacher}")
            print("Занятие успешно добавлено")
        except Exception as e:
            print(f"Ошибка: {e}")

    def delete_lesson(self):
        """Функция для удаления занятия из расписания. Вводятся группа и день недели. Создается пустой список,
        происходит чтение файла и поиск строки соответсвующих условию,
        которые записываются в список, из которого данные обратно перезапишутся, но без этой записи"""
        group = input("Введите группу: ")
        day = input("Введите день недели: ")

        with open('schedule.txt', 'r', encoding='utf-8') as file:
            lines = file.readlines()

        new_lines = []
        for line in lines:
            ls = line.strip().split(',')
            if ls[0] != group and ls[1] != day:
                new_lines.append(line)

        with open('schedule.txt', 'w', encoding='utf-8') as file:
            file.writelines(new_lines)

        print("Занятие удалено")

    def edit_lesson(self):
        """Функция для редактирования элемента расписания.
        Вводятся группа и день недели, номер элемента редактирования.
        В зависимости от выбора соответсвующий элемент меняется на введенный пользователем и обратно заспиысваются """
        group = input("Введите группу: ")
        day = input("Введите день недели: ")
        f = input(
            "Выберите соответствующий элемент редактирования: \n 1. День недели\n 2. Время\n 3. Преподаватель\n")
        with open('schedule.txt', 'r', encoding='utf-8') as file:
            lines = file.readlines()
            for i in range(len(lines)):
                ls = lines[i].strip().split(',')
                if ls[0] == group and ls[1] == day:
                    if f == '1':
                        new_day = input("Введите новый день проведения занятия: ")
                        ls[1] = new_day
                    elif f == '2':
                        new_time = input("Введите новое время занятия: ")
                        ls[2] = new_time
                    elif f == '3':
                        new_teacher = input("Введите нового преподавателя: ")
                        ls[4] = new_teacher
            else:
                print("В расписании нет данных для группы на этот день")

                lines[i] = ','.join(ls) + '\n'
        with open('schedule.txt', 'w', encoding='utf-8') as file:
            file.writelines(lines)
            print("Расписание успешно отредактировано")

    def edit_schedule(self):
        """Интерфейс для редактирования расписания"""
        f = input(
            "Выберите действие:\n 1. Добавление записи в расписание\n 2. Удаление записи из расписания\n 3. Изменение элемента расписания\n")
        if f == '1':
            self.new_lesson()
        elif f == '2':
            self.delete_lesson()
        elif f == '3':
            self.edit_lesson()

    def groups_info(self):
        """ Получает информацию о группе. аналогично выводу инфморации о пользователях
        вводится группа, если в фале groups есть такая группа, выводятся данные
        """
        group = input("Введите группу: ")
        with open('groups.txt', 'r', encoding='utf') as file:
            for line in file:
                st = line.strip().split(',')
                if st[0] == group:
                    group_info = Groups(st[0], st[1], st[2], st[3])
                    print(group_info)

    def add_group(self):
        """Функция для добавления новой группы. Вводятся данные, создается экземпляр с веденными данными.
        Данные записываются в файл с группами"""
        try:
            group = input("Введите группу: ")
            teacher = input("Введите преподавателя: ")
            level = input("Введите уровень: ")
            course = input("Введите курс: ")
            group = Groups(group, teacher, level, course)
            with open('groups.txt', 'a', encoding='utf-8') as file:
                file.write(f"{group.group_id},{group.teacher},{group.level},{group.course}\n")
            print("Группа успешно добавлена")
        except Exception as e:
            print(f"Ошибка: {e}")

    def students_to_excel(self):
        """Функция для выгрузки студентов в excel на текущий момент. Данные считываются с файла students"""
        try:
            with open('students.txt', 'r', encoding='utf-8') as file:
                lines = file.readlines()
            wb = Workbook()
            ws = wb.active
            today = date.today()
            ws['A1'] = f"Количество студентов на {today}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:H1')
            headers = ["Логин", "Группа", "Фамилия", "Имя", "Отчество", "E-mail", "Телефон", "Дата рождения"]
            ws.append(headers)
            for line in lines:
                st = line.strip().split(',')
                ws.append(st)
            wb.save('new_excel.xlsx')
            print(f"Данные экспортированы в файл new_excel.xlsx")
            os.startfile('new_excel.xlsx')
        except Exception as e:
            print(f"Ошибка: {e}")

def manager_interface():
    """ Интерфейс для взаимодействия с менеджером. В зависимости от ввода номера действия, вызывается соответствующая функция"""
    manager = Managers()
    while True:
        f = input(f"""Выберите действие: 
        1. Распечатать анкету для нового студента/преподавателя
        2. Регистрация нового преподавателя
        3. Регистрация нового студента
        4. Отслеживание расписания
        5. Редактирование расписания
        6. Личные данные преподавателя/ученика
        7. Распечатать сертификат
        8. Информация о группе
        9. Добавить группу
        10. Выгрузка в excel
        11. Выход из личного кабинета\n""")
        if f == '1':
            manager.print_file()
        elif f == '2':
            manager.register_teacher()
        elif f == '3':
            manager.register_student()
        elif f == '4':
            manager.get_schedule_for_group()
        elif f == '5':
            manager.edit_schedule()
        elif f == '6':
            manager.get_info()
        elif f == '7':
            student_name = input("Введите ФИО студента (в родительном падеже): ")
            language = input("Введите язык, который был изучен: ")
            level = input("Введите уровень языка студента: ")
            course = input("Введите курс: ")
            organization = input("Введите полное название организации: ")
            tdate = input("Дата выдачи сертификата: ")
            manager.print_certificate(student_name, language, level, course, organization, tdate)
        elif f == '8':
            manager.groups_info()
        elif f == '9':
            manager.add_group()
        elif f == '10':
            manager.students_to_excel()
        elif f == '11':
            break